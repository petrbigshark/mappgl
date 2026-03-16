#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import os
import re
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple, Dict

import openpyxl
from openpyxl.styles import Font


NEEDLE_PREFIX = "Отсутствует конфигурация для сезона с типом"
REQUIRED_ATTR_28_PREFIX = "Не задано значение обязательного атрибута 28"
# Example in reason: "Отсутствует конфигурация для сезона с типом '25SS' и годом 'null'; ..."
SEASON_BLOCK_PATTERNS = [
    re.compile(
        r"Отсутствует\s+конфигурация\s+для\s+сезона\s+с\s+типом\s*(.+?)(?:\s+и\s+годом\b|;|$)",
        flags=re.IGNORECASE,
    ),
    re.compile(
        r"Не\s*задано\s*значение\s*обязательного\s*атрибута\s*28\s*:\s*(.+?)(?:;|$)",
        flags=re.IGNORECASE,
    ),
]
SEASON_FALLBACK_TOKEN_RE = re.compile(r"\b(?:\d{2}(?:SS|AW|FW)|(?:SS|AW|FW)\d{2})\b", flags=re.IGNORECASE)
SEASON_CODE_PREFIX_YEAR_RE = re.compile(r"\b(FW|SS|AW|W|S)\s+(\d{2})\b", flags=re.IGNORECASE)
SEASON_CODE_SUFFIX_YEAR_RE = re.compile(r"\b(\d{2})\s+(FW|SS|AW|W|S)\b", flags=re.IGNORECASE)
SEASON_CODE_ONLY_RE = re.compile(r"^(?:FW|SS|AW|W|S)\d{2}$|^\d{2}(?:FW|SS|AW|W|S)$|^(?:FW|SS|AW|W|S)$", flags=re.IGNORECASE)

OUTPUT_HEADERS = [
    "id",
    "objectType",
    "expr",
    "outputValue",
    "matchType",
    "repeatMatching",
    "params",
    "indexNumber",
    "clientEmail",
]


def load_regex_mapping(mapping_xlsx: str) -> List[Tuple[re.Pattern, str]]:
    """
    mapping_xlsx (справочник) format:
      A: expr (regex as string)
      B: outputValue (AW/SS)
    First row is header.
    """
    wb = openpyxl.load_workbook(mapping_xlsx)
    ws = wb[wb.sheetnames[0]]

    pairs: List[Tuple[re.Pattern, str]] = []
    for r in range(2, ws.max_row + 1):
        expr = ws.cell(r, 1).value
        outv = ws.cell(r, 2).value
        if not expr or not outv:
            continue

        outv = str(outv).strip().upper()
        if outv not in {"AW", "SS"}:
            continue

        expr = str(expr).strip()
        try:
            pat = re.compile(expr, flags=re.IGNORECASE)
            pairs.append((pat, outv))
        except re.error:
            # bad regex -> skip
            continue
    return pairs


def extract_season_tokens_from_reason(reason: str) -> List[str]:
    """
    Extract season tokens from the semantic block after "...с типом" or
    from attribute 28 payload, then preserve the display form for expr.
    """
    if not reason:
        return []
    s = str(reason)
    s_cf = s.casefold()
    if not any(prefix.casefold() in s_cf for prefix in (NEEDLE_PREFIX, REQUIRED_ATTR_28_PREFIX)):
        return []

    def normalize_display_token(raw: str) -> str:
        token = str(raw or "").strip()
        token = token.replace("’", "'").replace("`", "'")
        token = re.sub(r"\s+", " ", token).strip()
        while token and token[0] in {"'", '"'}:
            token = token[1:].strip()
        while token and token[-1] in {"'", '"'}:
            token = token[:-1].strip()
        return token

    tokens: List[str] = []
    for pat in SEASON_BLOCK_PATTERNS:
        for m in pat.finditer(s):
            token = normalize_display_token(m.group(1))
            if token:
                tokens.append(token)

    if not tokens:
        for m in SEASON_FALLBACK_TOKEN_RE.finditer(s):
            token = normalize_display_token(m.group(0))
            if token:
                tokens.append(token)

    # De-duplicate while preserving order.
    seen = set()
    unique_tokens: List[str] = []
    for token in tokens:
        key = token.casefold()
        if key in seen:
            continue
        seen.add(key)
        unique_tokens.append(token)

    return unique_tokens


def normalize_lookup_token(token: str) -> str:
    value = str(token or "").strip()
    value = value.replace("’", "'").replace("`", "'")
    value = re.sub(r"[\"']+", "", value)
    value = re.sub(r"\s+", " ", value).strip()
    value = SEASON_CODE_PREFIX_YEAR_RE.sub(lambda m: f"{m.group(1).upper()}{m.group(2)}", value)
    value = SEASON_CODE_SUFFIX_YEAR_RE.sub(lambda m: f"{m.group(1)}{m.group(2).upper()}", value)
    if SEASON_CODE_ONLY_RE.fullmatch(value):
        return value.upper()
    return value


def map_by_reference(token: str, ref: List[Tuple[re.Pattern, str]]) -> Optional[str]:
    for pat, outv in ref:
        if pat.search(token):
            return outv
    return None


def map_by_llm(token: str, model: str = "gpt-5") -> Optional[str]:
    """
    Optional: requires OPENAI_API_KEY and 'openai' package installed.
    Must return strictly 'AW' or 'SS'.
    """
    try:
        from openai import OpenAI  # type: ignore
    except Exception:
        return None

    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return None

    client = OpenAI(api_key=api_key)
    prompt = (
        "Ты классификатор сезонного типа для fashion.\n"
        "Нужно сопоставить входной сезонный токен к одному из двух значений:\n"
        "- SS (Spring/Summer)\n"
        "- AW (Fall/Winter)\n\n"
        "Верни строго одно из: SS или AW.\n"
        f"Токен: {token}\n"
        "Ответ:"
    )

    try:
        resp = client.responses.create(
            model=model,
            input=prompt,
            temperature=0
        )
        text = (resp.output_text or "").strip().upper()
    except Exception:
        return None

    m = re.search(r"\b(SS|AW)\b", text)
    return m.group(1) if m else None


def get_versioned_output_dir(base_output: Path, email: str, date_str: str) -> Path:
    base_output.mkdir(parents=True, exist_ok=True)
    prefix = f"{email} {date_str} v"

    max_v = 0
    for p in base_output.iterdir():
        if not p.is_dir():
            continue
        name = p.name
        if not name.startswith(prefix):
            continue
        m = re.search(r"\bv(\d+)\b", name)
        if m:
            max_v = max(max_v, int(m.group(1)))

    new_dir = base_output / f"{email} {date_str} v{max_v + 1}"
    new_dir.mkdir(parents=True, exist_ok=False)
    return new_dir


def _sheet_key(name: str) -> str:
    return "".join(str(name or "").split()).casefold()


def normalize_reason_key(reason: Optional[str]) -> str:
    return " ".join(str(reason or "").split()).casefold()


def resolve_input_sheet_name(sheet_names: List[str], preferred_sheet: Optional[str] = "Result 1") -> str:
    if not sheet_names:
        raise ValueError("Во входном файле нет листов.")

    by_key = {_sheet_key(name): name for name in sheet_names}
    if preferred_sheet:
        hit = by_key.get(_sheet_key(preferred_sheet))
        if hit:
            return hit

    hit = by_key.get(_sheet_key("Result 1"))
    if hit:
        return hit
    return sheet_names[0]


def read_input_sheet(
    input_xlsx: str,
    preferred_sheet: Optional[str] = "Result 1",
) -> Tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet, Dict[str, int]]:
    wb = openpyxl.load_workbook(input_xlsx)
    resolved_sheet = resolve_input_sheet_name(wb.sheetnames, preferred_sheet)
    ws = wb[resolved_sheet]

    header_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h is None:
            continue
        header_map[str(h).strip().lower()] = c

    if "reason" not in header_map:
        raise ValueError("Во входном файле нет колонки 'reason' (заголовок в первой строке).")
    if "storecode" not in header_map:
        raise ValueError("Во входном файле нет колонки 'storecode' (заголовок в первой строке).")

    return wb, ws, header_map


def build_records_from_input(ws, header_map, ref_map, index_start: int, email: str,
                             use_llm: bool, llm_model: str):
    """
    Returns:
      records: list for upload file
      errors: list for errors file (storecode, token, reason)
    """
    records = []
    errors = []

    seen_pairs = set()  # (expr, outputValue) to avoid duplicates if needed
    seen_reason_keys = set()
    row_id = 1
    index_num = index_start

    reason_col = header_map["reason"]
    storecode_col = header_map["storecode"]

    for r in range(2, ws.max_row + 1):
        reason = ws.cell(r, reason_col).value
        storecode = ws.cell(r, storecode_col).value

        reason_key = normalize_reason_key("" if reason is None else str(reason))
        if not reason_key or reason_key in seen_reason_keys:
            continue
        seen_reason_keys.add(reason_key)

        tokens = extract_season_tokens_from_reason("" if reason is None else str(reason))
        if not tokens:
            continue

        for token in tokens:
            lookup_token = normalize_lookup_token(token)
            outv = map_by_reference(lookup_token, ref_map)
            if outv is None and use_llm:
                outv = map_by_llm(lookup_token, model=llm_model)

            if outv is None:
                errors.append((storecode, token, reason))
                continue

            key = (token, outv)
            if key in seen_pairs:
                continue
            seen_pairs.add(key)

            records.append((row_id, "season type", token, outv, index_num, email))
            row_id += 1
            index_num += 1

    return records, errors


def write_upload_file(path: Path, records):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"

    for c, h in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)

    for i, (rid, obj, expr, outv, index_num, email) in enumerate(records, start=2):
        ws.cell(i, 1, rid)
        ws.cell(i, 2, obj)
        ws.cell(i, 3, expr)
        ws.cell(i, 4, outv)
        ws.cell(i, 5, "EQUALS")
        ws.cell(i, 6, "false")
        ws.cell(i, 7, "{}")
        ws.cell(i, 8, index_num)
        ws.cell(i, 9, email)

    wb.save(path)


def write_errors_file(path: Path, errors):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"
    headers = ["storecode", "token", "reason"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)

    for i, (storecode, token, reason) in enumerate(errors, start=2):
        ws.cell(i, 1, storecode)
        ws.cell(i, 2, token)
        ws.cell(i, 3, "" if reason is None else str(reason))

    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Входящий Excel (как на скрине), с колонками storecode и reason")
    ap.add_argument(
        "--mapping",
        required=False,
        default=None,
        help="(Необязательно) путь к справочнику. По умолчанию берётся 'Справочник.xlsx' рядом с main.py",
    )
    ap.add_argument("--index-start", required=True, type=int, help="Стартовое значение indexNumber")
    ap.add_argument("--email", required=True, help="Почта клиента (пишется в clientEmail)")
    ap.add_argument("--output-dir", default="output", help="Корневая output-папка")
    ap.add_argument("--use-llm", action="store_true", help="Если включить — при промахе по справочнику спросим LLM")
    ap.add_argument("--llm-model", default="gpt-5", help="Модель для LLM (если включён --use-llm)")
    ap.add_argument("--sheet-name", default="Result 1", help="Предпочитаемый лист входного файла (fallback: первый лист)")

    args = ap.parse_args()

    # Справочник по умолчанию лежит рядом с main.py и называется "Справочник.xlsx"
    script_dir = Path(__file__).resolve().parent
    mapping_path = Path(args.mapping) if args.mapping else (script_dir / "Справочник.xlsx")
    if not mapping_path.exists():
        raise FileNotFoundError(
            f"Не найден справочник: {mapping_path}. Положи 'Справочник.xlsx' рядом с main.py "
            "или передай путь через --mapping."
        )

    ref_map = load_regex_mapping(str(mapping_path))
    wb, ws, header_map = read_input_sheet(args.input, preferred_sheet=args.sheet_name)

    records, errors = build_records_from_input(
        ws=ws,
        header_map=header_map,
        ref_map=ref_map,
        index_start=args.index_start,
        email=args.email,
        use_llm=bool(args.use_llm),
        llm_model=args.llm_model
    )

    if not records and not errors:
        print("Ошибок нет.")
        return
    
    date_str = datetime.now().strftime("%d.%m.%Y")
    out_root = Path(args.output_dir)
    out_dir = get_versioned_output_dir(out_root, args.email, date_str)

    upload_name = f"Сезоны для загрузки {args.email} {date_str}.xlsx"
    upload_path = out_dir / upload_name
    write_upload_file(upload_path, records)

    if errors:
        errors_path = out_dir / "ошибки сезонов.xlsx"
        write_errors_file(errors_path, errors)

    print(f"OK. Создано: {upload_path}")
    if errors:
        print(f"Есть ошибки: {out_dir / 'ошибки сезонов.xlsx'}")
    else:
        print("Ошибок нет.")


if __name__ == "__main__":
    main()
