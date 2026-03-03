#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Pattern, Tuple

import pandas as pd
import yaml
from openpyxl.styles import PatternFill
from unidecode import unidecode

from llm_client import LLMConfig, LLMError, SizeCategoryClassifier

DENIM_PANTS_LABEL = "Джинсовые размеры брюк"
DENIM_SKIRTS_LABEL = "Джинсовые размеры юбок"
DENIM_SLASH_LABEL = "Джинсовые размеры с /"
HIGHLIGHT_LIGHT_GREEN_HEX = "FFC6EFCE"


def die(msg: str, code: int = 2) -> None:
    print(f"ERROR: {msg}", file=sys.stderr)
    raise SystemExit(code)


def load_yaml(path: Path) -> Dict[str, Any]:
    if not path.exists():
        die(f"Config not found: {path}")
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def _sheet_key(name: Any) -> str:
    return "".join(str(name or "").split()).casefold()


def resolve_input_sheet(path: Path, preferred_sheet: Any = None) -> str:
    with pd.ExcelFile(path) as xls:
        sheet_names = list(xls.sheet_names)
    if not sheet_names:
        die(f"No sheets in input file: {path}")

    by_key = {_sheet_key(name): name for name in sheet_names}
    if isinstance(preferred_sheet, str) and preferred_sheet.strip():
        hit = by_key.get(_sheet_key(preferred_sheet))
        if hit:
            return hit

    for candidate in ("Result 1",):
        hit = by_key.get(_sheet_key(candidate))
        if hit:
            return hit

    if isinstance(preferred_sheet, int) and 0 <= preferred_sheet < len(sheet_names):
        return sheet_names[preferred_sheet]
    return sheet_names[0]


def normalize_space(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def normalize_for_contains(s: Any) -> str:
    return normalize_space(s).casefold()


def normalize_brand(s: Any) -> str:
    x = normalize_space(s)
    if not x:
        return ""
    x = x.replace("’", "'").replace("`", "'")
    x = x.replace("&", " and ")
    x = x.replace("@", "a")
    x = re.sub(r"[™®©]", "", x)
    x = unidecode(x)
    x = x.casefold()
    x = re.sub(r"[^a-z0-9\s'/\-]", " ", x)
    x = re.sub(r"[\-_/]", " ", x)
    x = re.sub(r"\s+", " ", x).strip()
    return x


def normalize_text_for_keywords(s: Any) -> str:
    x = normalize_space(s)
    if not x:
        return ""
    x = unidecode(x).casefold()
    x = re.sub(r"[^a-z0-9а-яё\s\-]", " ", x)
    x = re.sub(r"[\-_/]", " ", x)
    x = re.sub(r"\s+", " ", x).strip()
    return x


def resolve_column(df: pd.DataFrame, wanted: str) -> Optional[str]:
    w = normalize_space(wanted).casefold()
    if not w:
        return None

    for c in df.columns:
        if normalize_space(c).casefold() == w:
            return c
    for c in df.columns:
        if w in normalize_space(c).casefold():
            return c
    return None


def sort_by_reason_az(df: pd.DataFrame, reason_col: str) -> pd.DataFrame:
    if reason_col not in df.columns:
        return df.reset_index(drop=True)
    tmp = df.copy()
    tmp["_reason_sort_key"] = tmp[reason_col].astype(str).map(normalize_space).str.casefold()
    tmp = tmp.sort_values(by=["_reason_sort_key"], kind="stable").drop(columns=["_reason_sort_key"])
    return tmp.reset_index(drop=True)


def infer_denim_size_error_reason_label(reason: Any, category: Any, parentcategory: Any) -> Optional[str]:
    text = normalize_space(reason)
    low = text.casefold()
    cat = normalize_space(category).casefold()
    _parent = normalize_space(parentcategory).casefold()

    if re.search(r"не найден размер\s*'[^']*/[^']*'", text, flags=re.IGNORECASE):
        return DENIM_SLASH_LABEL

    if "не найден тип размера" not in low:
        return None

    has_jeans = any(m in low for m in ("джинс", "jeans")) or any(m in cat for m in ("джинс", "jeans"))

    has_pants = any(m in low for m in ("брюк", "брюки", "pantaloni", "pants", "trousers")) or any(
        m in cat for m in ("брюк", "брюки", "pantaloni", "pants", "trousers")
    )
    if has_pants and not has_jeans:
        return DENIM_PANTS_LABEL

    has_skirts = any(m in low for m in ("юбк", "gonne", "skirt")) or any(m in cat for m in ("юбк", "gonne", "skirt"))
    if has_skirts:
        return DENIM_SKIRTS_LABEL

    return None


def build_size_errors_sheet(
    df_sorted: pd.DataFrame,
    reason_col: str,
    category_col: str,
    parent_col: str,
    base_cols: List[str],
    count_col: str = "Количество ошибок",
    reason_label_col: str = "Причина ошибок",
) -> pd.DataFrame:
    cols = list(base_cols) + [count_col, reason_label_col]
    if df_sorted.empty:
        return pd.DataFrame(columns=cols)

    work = df_sorted.copy()
    work["_error_label"] = work.apply(
        lambda r: infer_denim_size_error_reason_label(r.get(reason_col), r.get(category_col), r.get(parent_col)),
        axis=1,
    )
    work = work[work["_error_label"].notna()].copy()
    if work.empty:
        return pd.DataFrame(columns=cols)

    counts = work["_error_label"].value_counts(dropna=False)

    summary = work.drop_duplicates(subset=["_error_label"], keep="first").copy()
    summary[count_col] = summary["_error_label"].map(counts).astype(int)
    summary[reason_label_col] = summary["_error_label"]
    summary = summary[base_cols + [count_col, reason_label_col]]
    return summary.reset_index(drop=True)


def find_denim_highlight_row_indexes(
    df_sorted: pd.DataFrame,
    reason_col: str,
    category_col: str,
    parent_col: str,
) -> List[int]:
    if df_sorted.empty:
        return []
    labels = df_sorted.apply(
        lambda r: infer_denim_size_error_reason_label(r.get(reason_col), r.get(category_col), r.get(parent_col)),
        axis=1,
    )
    return [int(i) for i in labels[labels.notna()].index.tolist()]


def highlight_rows_in_sheet(ws: Any, row_indexes: List[int], max_col: int) -> None:
    if not row_indexes:
        return
    fill = PatternFill(start_color=HIGHLIGHT_LIGHT_GREEN_HEX, end_color=HIGHLIGHT_LIGHT_GREEN_HEX, fill_type="solid")
    for idx in row_indexes:
        excel_row = int(idx) + 2  # +1 header, +1 because idx is 0-based
        for col in range(1, max_col + 1):
            ws.cell(row=excel_row, column=col).fill = fill


def columns_for_errors_sheet(
    df_out: pd.DataFrame,
    size_category_col: str,
    size_type_col: str,
) -> List[str]:
    drop_cols = {size_category_col, size_type_col, "_unified_row_id"}
    base = [c for c in df_out.columns if c not in drop_cols]

    desc_col = resolve_column(df_out, "description")
    if desc_col and desc_col in base:
        i = base.index(desc_col)
        return base[: i + 1]
    return base


def resolve_reference_path(
    cli_value: Optional[str],
    config_filename: str,
    script_dir: Path,
    input_path: Path,
) -> Path:
    if cli_value:
        return Path(cli_value)

    candidates = [
        script_dir / config_filename,
        input_path.parent / config_filename,
        Path(config_filename),
    ]
    for c in candidates:
        if c.exists():
            return c

    # Return first candidate for clear error message downstream.
    return candidates[0]


def compile_reason_patterns(patterns: List[str]) -> List[Pattern[str]]:
    out: List[Pattern[str]] = []
    for p in patterns:
        try:
            out.append(re.compile(p, flags=re.IGNORECASE))
        except re.error as e:
            die(f"Bad regex in filters.reason_start_regex: {p}. Error: {e}")
    return out


def reason_starts_with_target(reason: Any, patterns: List[Pattern[str]]) -> bool:
    s = str(reason or "")
    return any(p.search(s) is not None for p in patterns)


def contains_any_marker(text: Any, markers: List[str]) -> bool:
    s = normalize_for_contains(text)
    if not s:
        return False
    return any(normalize_for_contains(m) in s for m in markers if normalize_space(m))


def parent_is_allowed(parentcategory: Any, include_markers: List[str], exclude_markers: List[str]) -> bool:
    if contains_any_marker(parentcategory, exclude_markers):
        return False
    if not include_markers:
        # Exclude-only mode: everything non-kids is allowed.
        return True
    return contains_any_marker(parentcategory, include_markers)


def heuristic_category_label(
    category: str,
    shoes_keywords: List[str],
    clothing_keywords: List[str],
    default_label: str,
) -> str:
    cat = normalize_text_for_keywords(category)
    if not cat:
        return default_label

    for kw in shoes_keywords:
        nkw = normalize_text_for_keywords(kw)
        if nkw and nkw in cat:
            return "SHOES"

    for kw in clothing_keywords:
        nkw = normalize_text_for_keywords(kw)
        if nkw and nkw in cat:
            return "CLOTHING"

    return default_label


def load_size_reference(path: Path) -> Tuple[Dict[str, str], List[str]]:
    if not path.exists():
        die(f"Reference file not found: {path}")

    df = pd.read_excel(path)
    c_brand = resolve_column(df, "brand")
    c_sizetype = resolve_column(df, "sizetype")
    if not c_brand or not c_sizetype:
        die(f"{path.name} must contain columns brand and sizetype. Found: {list(df.columns)}")

    mapping: Dict[str, str] = {}
    conflicts: List[str] = []

    for _, r in df.iterrows():
        brand_raw = normalize_space(r[c_brand])
        sizetype_raw = normalize_space(r[c_sizetype])
        if not brand_raw:
            continue

        key = normalize_brand(brand_raw)
        if not key:
            continue

        prev = mapping.get(key)
        if prev is None:
            mapping[key] = sizetype_raw
            continue

        if prev != sizetype_raw:
            conflicts.append(f"{brand_raw}: '{prev}' vs '{sizetype_raw}'")

    return mapping, conflicts


def ensure_output_run_dir(output_root: Path, input_stem: str) -> Tuple[Path, str, int]:
    output_root.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime("%d.%m.%Y")
    base = f"{input_stem} {date_str}"

    versions: List[int] = []
    for p in output_root.iterdir():
        if not p.is_dir():
            continue
        if not p.name.startswith(base):
            continue
        m = re.search(r"\sv(\d+)$", p.name)
        if m:
            versions.append(int(m.group(1)))

    v = (max(versions) + 1) if versions else 1
    run_dir = output_root / f"{base} v{v}"
    run_dir.mkdir(parents=True, exist_ok=False)
    return run_dir, date_str, v


def classify_categories(
    categories: List[str],
    cfg: Dict[str, Any],
    use_llm: bool,
) -> Tuple[Dict[str, str], Dict[str, str], Optional[str]]:
    cls_cfg = cfg["classification"]
    allowed_labels = {str(x).upper().strip() for x in cls_cfg.get("allowed_labels", [])}
    if not allowed_labels:
        allowed_labels = {"CLOTHING", "SHOES", "OTHER"}

    default_label = str(cls_cfg.get("default_label", "OTHER")).upper().strip()
    if default_label not in allowed_labels:
        default_label = "OTHER"

    shoes_keywords = cls_cfg.get("shoes_keywords", [])
    clothing_keywords = cls_cfg.get("clothing_keywords", [])

    llm_map: Dict[str, str] = {}
    llm_error: Optional[str] = None

    if use_llm and categories:
        llm_cfg = LLMConfig(
            model=str(cls_cfg.get("llm_model", "gpt-5-mini")),
            timeout_sec=int(cls_cfg.get("llm_timeout_sec", 90)),
            max_retries=int(cls_cfg.get("llm_max_retries", 3)),
        )
        classifier = SizeCategoryClassifier(llm_cfg)
        try:
            llm_map = classifier.classify(categories)
        except LLMError as e:
            llm_error = str(e)

    out_label: Dict[str, str] = {}
    out_source: Dict[str, str] = {}
    for cat in categories:
        label = str(llm_map.get(cat, "")).upper().strip()
        if label in allowed_labels:
            out_label[cat] = label
            out_source[cat] = "llm"
            continue

        h = heuristic_category_label(cat, shoes_keywords, clothing_keywords, default_label)
        if h not in allowed_labels:
            h = default_label
        out_label[cat] = h
        out_source[cat] = "heuristic"

    return out_label, out_source, llm_error


def main() -> None:
    parser = argparse.ArgumentParser(description="Oskelly Size Mapping")
    parser.add_argument("--input", required=True, help="Input Excel file with columns reason/brand/parentcategory/category")
    parser.add_argument("--email", default=None, help="Email for output filename (used by unified runner)")
    parser.add_argument("--clothing-dict", default=None, help="Excel dictionary for clothing: columns brand,sizetype")
    parser.add_argument("--shoes-dict", default=None, help="Excel dictionary for shoes: columns brand,sizetype")
    parser.add_argument("--config", default="config.yml", help="Path to config.yml")
    parser.add_argument("--outdir", default=None, help="Output root folder (overrides config)")
    parser.add_argument("--sheet-name", default=None, help="Input sheet name override")
    parser.add_argument("--dry-run", action="store_true", help="Do not call LLM, use heuristic category classification only")
    parser.add_argument("--disable-llm", action="store_true", help="Force-disable LLM and use heuristic category classification")
    args = parser.parse_args()

    cfg = load_yaml(Path(args.config))

    in_path = Path(args.input)
    if not in_path.exists():
        die(f"Input file not found: {in_path}")

    script_dir = Path(__file__).resolve().parent
    refs_cfg = cfg.get("references", {})
    clothing_dict_name = str(refs_cfg.get("clothing_dict_xlsx", "Справочник одежда.xlsx"))
    shoes_dict_name = str(refs_cfg.get("shoes_dict_xlsx", "Справочник обувь.xlsx"))
    clothing_dict_path = resolve_reference_path(args.clothing_dict, clothing_dict_name, script_dir, in_path)
    shoes_dict_path = resolve_reference_path(args.shoes_dict, shoes_dict_name, script_dir, in_path)

    preferred_sheet: Any = args.sheet_name if args.sheet_name is not None else cfg["input"].get("sheet_name", "Result 1")
    sheet_name = resolve_input_sheet(in_path, preferred_sheet)
    df_in = pd.read_excel(in_path, sheet_name=sheet_name)

    input_cols = cfg["input"]["columns"]
    c_reason = resolve_column(df_in, input_cols["reason"])
    c_brand = resolve_column(df_in, input_cols["brand"])
    c_parent = resolve_column(df_in, input_cols["parentcategory"])
    c_category = resolve_column(df_in, input_cols["category"])

    missing = [x for x in [("reason", c_reason), ("brand", c_brand), ("parentcategory", c_parent), ("category", c_category)] if not x[1]]
    if missing:
        miss = ", ".join(k for k, _ in missing)
        die(f"Could not resolve columns: {miss}. Found columns: {list(df_in.columns)}")

    reason_patterns = compile_reason_patterns(cfg["filters"]["reason_start_regex"])
    include_parent_markers = cfg["filters"].get("include_parent_markers", [])
    exclude_parent_markers = cfg["filters"].get("exclude_parent_markers", [])

    mask_reason = df_in[c_reason].astype(str).apply(lambda x: reason_starts_with_target(x, reason_patterns))
    df_reason = df_in[mask_reason].copy()

    mask_parent = df_reason[c_parent].astype(str).apply(
        lambda x: parent_is_allowed(x, include_parent_markers, exclude_parent_markers)
    )
    df_filtered = df_reason[mask_parent].copy()
    df_filtered["_category_clean"] = df_filtered[c_category].astype(str).map(normalize_space)

    unique_categories = sorted([x for x in df_filtered["_category_clean"].dropna().unique().tolist() if normalize_space(x)])

    can_use_llm = bool(
        cfg["classification"].get("llm_enabled", True)
        and not args.dry_run
        and not args.disable_llm
        and os.getenv("OPENAI_API_KEY")
    )

    llm_skip_reason: Optional[str] = None
    if not can_use_llm:
        if args.dry_run:
            llm_skip_reason = "dry_run"
        elif args.disable_llm:
            llm_skip_reason = "disabled_by_flag"
        elif not cfg["classification"].get("llm_enabled", True):
            llm_skip_reason = "disabled_in_config"
        elif not os.getenv("OPENAI_API_KEY"):
            llm_skip_reason = "OPENAI_API_KEY is not set"
        else:
            llm_skip_reason = "unknown"

    category_label_map, category_source_map, llm_error = classify_categories(unique_categories, cfg, can_use_llm)

    out_cols_cfg = cfg["output"]["add_columns"]
    col_size_category = str(out_cols_cfg.get("size_category", "sizeCategory"))
    col_size_type_mapped = str(out_cols_cfg.get("size_type_mapped", "sizeTypeMapped"))

    default_label = str(cfg["classification"].get("default_label", "OTHER")).upper().strip()
    df_filtered[col_size_category] = df_filtered["_category_clean"].map(category_label_map).fillna(default_label)

    keep_labels = {"CLOTHING", "SHOES", "OTHER"}
    df_target = df_filtered[df_filtered[col_size_category].isin(keep_labels)].copy()

    clothing_map, clothing_conflicts = load_size_reference(clothing_dict_path)
    shoes_map, shoes_conflicts = load_size_reference(shoes_dict_path)

    df_target["_brand_key"] = df_target[c_brand].astype(str).map(normalize_brand)
    df_target[col_size_type_mapped] = ""

    mask_clothing = df_target[col_size_category] == "CLOTHING"
    mask_shoes = df_target[col_size_category] == "SHOES"
    mask_other = df_target[col_size_category] == "OTHER"
    df_target.loc[mask_clothing, col_size_type_mapped] = (
        df_target.loc[mask_clothing, "_brand_key"].map(clothing_map).fillna("")
    )
    df_target.loc[mask_shoes, col_size_type_mapped] = (
        df_target.loc[mask_shoes, "_brand_key"].map(shoes_map).fillna("")
    )
    df_target.loc[mask_other, col_size_type_mapped] = "OTHER"

    if bool(cfg["output"].get("keep_only_mapped_size_type", False)):
        df_target = df_target[df_target[col_size_type_mapped].astype(str).str.strip() != ""].copy()

    unmatched = df_target[df_target[col_size_type_mapped].astype(str).str.strip() == ""].copy()
    unmatched_unique = (
        unmatched[[c_brand, c_category, col_size_category]]
        .astype(str)
        .drop_duplicates()
        .reset_index(drop=True)
    )

    df_out = df_target.drop(columns=["_category_clean", "_brand_key"], errors="ignore")
    df_out = sort_by_reason_az(df_out, c_reason)

    out_root = Path(args.outdir) if args.outdir else Path(cfg["output"].get("outdir", "output"))
    run_dir, date_str, version = ensure_output_run_dir(out_root, in_path.stem)

    name_key = normalize_space(args.email) if args.email else ""
    if not name_key:
        name_key = in_path.stem

    result_name_tmpl = str(cfg["output"].get("result_filename_template", "Размеры для маппинга {name_key} {date}.xlsx"))
    result_name = result_name_tmpl.format(
        input_stem=in_path.stem,
        email=(args.email or ""),
        name_key=name_key,
        date=date_str,
    )
    result_path = run_dir / result_name
    result_sheet_name = str(cfg["output"].get("result_sheet_name", "Sheet1"))
    errors_sheet_name = str(cfg["output"].get("errors_sheet_name", "Ошибки"))
    errors_count_col = str(cfg["output"].get("errors_count_column", "Количество ошибок"))
    errors_reason_col = str(cfg["output"].get("errors_reason_column", "Причина ошибок"))
    result_file_created = False
    errors_sheet_rows = 0
    highlighted_rows_count = 0

    if not df_out.empty:
        highlight_row_indexes = find_denim_highlight_row_indexes(df_out, c_reason, c_category, c_parent)
        highlighted_rows_count = int(len(highlight_row_indexes))

        error_base_cols = columns_for_errors_sheet(df_out, col_size_category, col_size_type_mapped)
        errors_sheet = build_size_errors_sheet(
            df_sorted=df_out,
            reason_col=c_reason,
            category_col=c_category,
            parent_col=c_parent,
            base_cols=error_base_cols,
            count_col=errors_count_col,
            reason_label_col=errors_reason_col,
        )
        errors_sheet_rows = int(len(errors_sheet))

        with pd.ExcelWriter(result_path, engine="openpyxl") as writer:
            df_out.to_excel(writer, index=False, sheet_name=result_sheet_name)
            if not errors_sheet.empty:
                errors_sheet.to_excel(writer, index=False, sheet_name=errors_sheet_name)
            ws_main = writer.book[result_sheet_name]
            highlight_rows_in_sheet(ws_main, highlight_row_indexes, max_col=len(df_out.columns))
        result_file_created = True

    category_map_name = str(cfg["output"].get("category_map_filename", "category_mapping.xlsx"))
    df_cat_map = pd.DataFrame(
        [
            {"category": cat, "label": category_label_map.get(cat, default_label), "source": category_source_map.get(cat, "heuristic")}
            for cat in unique_categories
        ]
    )
    df_cat_map.to_excel(run_dir / category_map_name, index=False)

    if not unmatched_unique.empty:
        unmatched_name = str(cfg["output"].get("unmatched_brands_filename", "unmatched_brands.xlsx"))
        unmatched_unique.to_excel(run_dir / unmatched_name, index=False)

    source_counts: Dict[str, int] = {}
    for v in category_source_map.values():
        source_counts[v] = source_counts.get(v, 0) + 1

    label_counts: Dict[str, int] = {}
    for v in category_label_map.values():
        label_counts[v] = label_counts.get(v, 0) + 1

    report = {
        "args": {
            "input": str(in_path),
            "email": args.email,
            "clothing_dict": str(clothing_dict_path),
            "shoes_dict": str(shoes_dict_path),
            "sheet_name": sheet_name,
            "dry_run": bool(args.dry_run),
            "disable_llm": bool(args.disable_llm),
        },
        "run": {
            "run_dir": str(run_dir),
            "date": date_str,
            "version": version,
            "result_file": str(result_path) if result_file_created else None,
            "result_file_created": bool(result_file_created),
            "errors_sheet_rows": int(errors_sheet_rows),
            "highlighted_rows_count": int(highlighted_rows_count),
        },
        "counts": {
            "input_rows": int(len(df_in)),
            "rows_after_reason_filter": int(len(df_reason)),
            "rows_after_parent_filter": int(len(df_filtered)),
            "unique_categories_for_classification": int(len(unique_categories)),
            "rows_after_category_filter": int(len(df_target)),
            "rows_with_empty_size_type": int(len(unmatched)),
            "rows_final_output": int(len(df_out)),
        },
        "classification": {
            "llm_used": bool(can_use_llm),
            "llm_skip_reason": llm_skip_reason,
            "llm_error": llm_error,
            "category_label_counts": label_counts,
            "category_source_counts": source_counts,
        },
        "references": {
            "clothing_brands": len(clothing_map),
            "shoes_brands": len(shoes_map),
            "clothing_conflicts": clothing_conflicts[:50],
            "shoes_conflicts": shoes_conflicts[:50],
        },
    }

    report_name = str(cfg["output"].get("report_filename", "run_report.json"))
    (run_dir / report_name).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print("OK")
    if result_file_created:
        print(f"Saved result: {result_path}")
    else:
        print("No rows for size mapping: result Excel was not created.")
    print(f"Saved report: {run_dir / report_name}")
    print(f"Rows: {len(df_in)} -> {len(df_out)}")
    if llm_error:
        print(f"LLM warning: {llm_error}")


if __name__ == "__main__":
    main()
